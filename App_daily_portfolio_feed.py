import os
import copy
import numpy as np
import pandas as pd
#import xlwings as xlw
import Lib_Market_Akit  as IAL_App
import Config_Rating_Mapping as Rating_Cofig
import Config_BSCR as BSCR_Cofig
import datetime
from openpyxl import load_workbook
from pandas.tseries.offsets import MonthEnd

akit_dir = 'C:/AKit v4.1.0/BIN'
os.sys.path.append(akit_dir)
import MktPython3 as MKT
import IALPython3 as IAL
import math

### Kyle:
### Inputs of actual_portfolio_feed is changed but Inputs of daily_portfolio_feed is not updated

def daily_portfolio_feed(eval_date, valDate_base, workDir, fileName, asset_fileName_T_plus_1, Price_Date, market_factor, output = 1, mappingFile = '.\Mapping.xlsx', ratingMapFile = '.\Rating_Mapping.xlsx'):
    def wavg(val_col_name, wt_col_name):
        def inner(group):
            return (group[val_col_name] * group[wt_col_name]).sum() / group[wt_col_name].sum()

        inner.__name__ = 'wtd_avg'
        return inner

#    workDir = r'L:\\DSA Re\\Workspace\\Production\\EBS Dashboard\\Python_Code\\Asset_Holding_Feed'
    os.chdir(workDir)

#    fileName = r'.\DSA RE Holdings.xlsx'

    portFile = pd.ExcelFile(fileName)
#    portInput = pd.read_excel(portFile, sheet_name='DSA RE Holdings', skiprows=[0, 1, 2, 3, 4, 5, 6])
    portInput = pd.read_excel(portFile, sheet_name='AssetSummaryFromPython')

    portInput = portInput.dropna(axis=0, how='all')
    portInput = portInput.dropna(axis=1, how='all')
    # portInput = portInput.fillna('')
  
    leMapFile = pd.ExcelFile(mappingFile)
    leMap = leMapFile.parse('LegalEntity')
    portInput = portInput.merge(leMap, how='left', left_on=['Owning Entity Name', 'Owning Entity GEMS ID'],
                                right_on=['Owning Entity Name', 'Owning Entity GEMS ID'])

    ratingMapFile = pd.ExcelFile(ratingMapFile)
    ratingMap = ratingMapFile.parse('naic')

    portInput = portInput.merge(ratingMap, how='left', left_on=['NAIC Rating', 'AIG Derived Rating'],
                                right_on=['NAIC Rating', 'AIG Derived Rating'])

    # Two CUSIPs are missing in the asset summary as their 'AIG Asset Class 3' is N/A.
    portInput.loc[portInput['Lot Number Unique ID'] == '218521143_121643527_1951603', ['AIG Asset Class 3']] = 'Corporate Bond'
    portInput.loc[portInput['Lot Number Unique ID'] == '218521146_121464874_1951604', ['AIG Asset Class 3']] = 'Corporate Bond'
    
    # Split out ML-III B Notes
    portInput['AIG Asset Class 3'] = np.where((portInput['Issuer Name'] == 'LSTREET II, LLC'), "ML-III B-Notes",
                                              portInput['AIG Asset Class 3'])

    # Default setting NA 'AIG Derived Rating' to 'BBB'
    portInput['AIG Derived Rating'].fillna('NA-BBB', inplace=True)

    # Handling surplus account
    portInput['SurplusAccount'] = 'Not-Surplus'

    portInput['SurplusAccount'] = np.where(
        ((portInput['Category'] == "Surplus") & (portInput['Strategy (Hld) Long Desc'] == "SAM RE OVERSEAS POOL LIFE")),
        "Long Term Surplus", portInput['SurplusAccount'])

    portInput['SurplusAccount'] = np.where(
        ((portInput['Category'] == "Surplus") & (portInput['Strategy (Hld) Long Desc'] == "SAM RE OVERSEAS POOL PC")),
        "General Surplus", portInput['SurplusAccount'])

    portInput['SurplusAccount'] = np.where(
        ((portInput['SurplusAccount'] == 'Not-Surplus') &
         (portInput['Category'] == "Surplus") & (portInput['Encumbrance Program Level 4 Desc'] == 'DSA REINSURANCE PC')),
        "General Surplus", portInput['SurplusAccount'])

    portInput['SurplusAccount'] = np.where(
        ((portInput['SurplusAccount'] == 'Not-Surplus') &
         (portInput['Category'] == "Surplus") & (portInput['Encumbrance Program Level 4 Desc'] == 'DSA REINSURANCE LR')),
        "Long Term Surplus", portInput['SurplusAccount'])

    portInput['SurplusAccount'] = np.where(
        ((portInput['SurplusAccount'] == 'Not-Surplus') &
         (portInput['Category'] == "Surplus") & (
                 portInput['Encumbrance Program Level 4 Desc'] == 'DSA RE/AGL Trust- DSA REINSURANCE LR')),
        "Long Term Surplus", portInput['SurplusAccount'])

    portInput['SurplusAccount'] = np.where(
        ((portInput['SurplusAccount'] == 'Not-Surplus') &
         (portInput['Category'] == "Surplus") & (
                 portInput['Encumbrance Program Level 4 Desc'] == 'DSA RE/USL Trust - DSA REINSURANCE LR')),
        "Long Term Surplus", portInput['SurplusAccount'])

    portInput['SurplusAccount'] = np.where(
        ((portInput['SurplusAccount'] == 'Not-Surplus') &
         (portInput['Category'] == "Surplus") & (
                 portInput['Encumbrance Program Level 4 Desc'] == 'DSA RE/VALIC Trust - DSA REINSURANCE LR')),
        "Long Term Surplus", portInput['SurplusAccount'])

    portInput['Category'] = np.where(
        ((portInput['SurplusAccount'] == 'Not-Surplus') &
         (portInput['Category'] == "Surplus") & (portInput['Encumbrance Program Level 4 Desc'] == 'SOURCE UNDEFINED')),
        "ModCo", portInput['Category'])

    portInput.dropna(axis=0, how='all', subset=['Category'], inplace=True)

    portInput['Category'] = np.where((portInput['SurplusAccount'] != 'Not-Surplus'),
                                     portInput['SurplusAccount'], portInput['Category'])

    # clean up fake booking due to accounting system migration
    portInput = portInput[[not v.startswith('F-') for (i, v) in portInput['Portfolio (Source) Long Name'].iteritems()]]


    # Report ModCo Currency derivatives
    Cur_Der = portInput.loc[(portInput['AIG Asset Class 1'] == 'Derivative') & (portInput['Owning Entity Name'] != 'American International Reinsurance Company, Ltd.' ) & (portInput['Security Desc DESC'].str[:2] == 'FX'), "Market Value USD GAAP"].sum()
    ModCo_Non_IR_Der = portInput.loc[((portInput['AIG Asset Class 1'] == 'Derivative') & (portInput['Category'] == 'ModCo' )) & ((portInput['Security Desc DESC'].str[:8] != 'Interest')), "Market Value USD GAAP"].sum()    
    IR_Der = portInput.loc[(portInput['AIG Asset Class 1'] == 'Derivative') & (portInput['Owning Entity Name'] != 'American International Reinsurance Company, Ltd.' ) & (portInput['Security Desc DESC'].str[:8] == 'Interest'), "Market Value USD GAAP"].sum()
    LPT_Non_IR_Der = portInput.loc[((portInput['AIG Asset Class 1'] == 'Derivative') & (portInput['Category'] == 'LPT' )) , "Market Value USD GAAP"].sum()    
    
    print('Currency_Derivative_' + eval_date.strftime('%Y%m%d') + ': ' + str(Cur_Der))
    print('IR_Derivative_' + eval_date.strftime('%Y%m%d') + ': ' + str(IR_Der))

#### adjust Price Date that is zero in the IDR
#    portInput['Price Date'] = portInput['Price Date'].fillna(valDate_base,inplace = True)
#    portInput['Price Date'] = portInput['Price Date'].fillna(0)
#    portInput['Price Date'] = np.where(
#        ((portInput['Price Date'] == "00:00:00")|(portInput['Price Date'] == 0)), valDate_base , portInput['Price Date']) 
    
    
#    if eval_date > datetime.datetime(2019, 8, 31):
##        # Illiquidity impact estimation
#        price_date_0 = eval_date + MonthEnd(-2)
#        price_date_1 = eval_date + MonthEnd(-1)
#        print(type(portInput['Price Date']))
#        print(type(price_date_1))
#        Yield_Change_0 = market_factor[market_factor['val_date'] == eval_date]['IR'].values[0] - market_factor[market_factor['val_date'] == price_date_0]['IR'].values[0] \
#                       + 0.5 * (market_factor[market_factor['val_date'] == eval_date]['Credit_Spread'].values[0]/10000 - market_factor[market_factor['val_date'] == price_date_0]['Credit_Spread'].values[0]/10000)
#                       
#        Yield_Change_1 = market_factor[market_factor['val_date'] == eval_date]['IR'].values[0] - market_factor[market_factor['val_date'] == price_date_1]['IR'].values[0] \
#                       + 0.5 * (market_factor[market_factor['val_date'] == eval_date]['Credit_Spread'].values[0]/10000 - market_factor[market_factor['val_date'] == price_date_1]['Credit_Spread'].values[0]/10000)
#
#        Initial_mv_acc_int = portInput.loc[(portInput['Price Date'] <= price_date_1) & (portInput['AIG Asset Class 1'] == 'Fixed Income') & (portInput['Effective Duration (WAMV)'].notnull() ) & (portInput['Effective Duration (WAMV)'] != 0), 'Market Value with Accrued Int USD GAAP'].sum()
#            
#        portInput.loc[(portInput['Price Date'] > price_date_0) & (portInput['Price Date'] <= price_date_1) & (portInput['AIG Asset Class 1'] == 'Fixed Income') & (portInput['Effective Duration (WAMV)'].notnull() ) & (portInput['Effective Duration (WAMV)'] != 0) & (portInput['AIG Asset Class 3'] != 'ML-III B-Notes'), "Market Value with Accrued Int USD GAAP"] = portInput['Market Value with Accrued Int USD GAAP'] * (1 + portInput['Effective Duration (WAMV)'] * -Yield_Change_1) 
#        portInput.loc[(portInput['Price Date'] <= price_date_0) & (portInput['AIG Asset Class 1'] == 'Fixed Income') & (portInput['Effective Duration (WAMV)'].notnull() ) & (portInput['Effective Duration (WAMV)'] != 0)& (portInput['AIG Asset Class 3'] != 'ML-III B-Notes'), "Market Value with Accrued Int USD GAAP"] = portInput['Market Value with Accrued Int USD GAAP'] * (1 + portInput['Effective Duration (WAMV)'] * -Yield_Change_0) 
#                    
#        total_illiquid_adj = portInput.loc[(portInput['Price Date'] <= price_date_1) & (portInput['AIG Asset Class 1'] == 'Fixed Income') & (portInput['Effective Duration (WAMV)'].notnull() ) & (portInput['Effective Duration (WAMV)'] != 0), 'Market Value with Accrued Int USD GAAP'].sum() - Initial_mv_acc_int
#       
#        portInput.fillna(0, inplace=True)
#        portInput['Market Value USD GAAP'] = portInput['Market Value with Accrued Int USD GAAP'] - portInput['Accrued Int USD GAAP']
#        
#        print(eval_date.strftime('%Y%m%d') + 'Yield_Change_' + price_date_0.strftime('%Y%m%d') + ': ' + str(Yield_Change_0))
#        print(eval_date.strftime('%Y%m%d') + 'Yield_Change_' + price_date_1.strftime('%Y%m%d') + ': ' + str(Yield_Change_1))
#        print(eval_date.strftime('%Y%m%d') + '_Initial_mv_acc_int: ' + str(Initial_mv_acc_int))
#        print(eval_date.strftime('%Y%m%d') + '_total_illiquid_adj: ' + str(total_illiquid_adj))
        
    ### === Calculating Risk Charge for FI and Equity === ###
    RaMap_sp = pd.read_excel(leMapFile, sheet_name ='Rating_SP')
    RaMap_moodys = pd.read_excel(leMapFile, sheet_name ='Rating_Moodys')
    RaMap_fitch = pd.read_excel(leMapFile, sheet_name ='Rating_Fitch')
    RaMap_aig = pd.read_excel(leMapFile, sheet_name ='Rating_AIG')
    AsMap = pd.read_excel(leMapFile, sheet_name ='AssetClass' )
    RcMap = pd.read_excel(leMapFile, sheet_name = 'Asset_Risk_Charge')


    portInput['Fort Re Corp Segment'] = portInput['Category']

    # clean up fake booking due to accounting system migration
#    portInput = portInput[[not v.startswith('F-') for (i, v) in portInput['Portfolio (Source) Long Name'].iteritems()]]
    
    portInput['MV_USD_GAAP'] = portInput['Market Value USD GAAP']
    # zero out Non ALBA derivatives
    portInput['Market Value USD GAAP'] = np.where(
            ((portInput['AIG Asset Class 1'] =='Derivative') &(portInput['Category'] != 'ALBA')) ,0, portInput['Market Value USD GAAP'])
    
    portInput['Market Value with Accrued Int USD GAAP'] = np.where(
            ((portInput['AIG Asset Class 1'] =='Derivative') &(portInput['Category'] != 'ALBA')) ,0, portInput['Market Value with Accrued Int USD GAAP'])
    
    
    # Assign number scale to ratings
    portInput = portInput.merge(RaMap_sp, how = 'left', left_on = 'S&P Rating', right_on = 'S&P')\
                        .rename(columns = {'BSCR rating_SP': 'S&P_num'})\
                        .drop(columns = 'S&P')
 
    portInput = portInput.merge(RaMap_moodys, how = 'left', left_on = "Moody's Rating", right_on = "Moody's")\
                        .rename(columns = {'BSCR rating_Moodys': 'Moody_num'})\
                        .drop(columns = "Moody's")
                        
    portInput = portInput.merge(RaMap_fitch, how = 'left', left_on = 'Fitch Rating', right_on = 'Fitch')\
                        .rename(columns = {'BSCR rating_Fitch': 'Fitch_num'})\
                        .drop(columns = 'Fitch')

    portInput = portInput.merge(RaMap_aig, how = 'left', left_on = 'AIG Derived Rating', right_on = 'AIG Rating')\
                        .rename(columns = {'BSCR rating_AIG': 'AIG_num'})\
                        .drop(columns = 'AIG Rating')

    # Assign BSCR rating
    portInput['Mapped_BSCR_Rating'] = 0
   # ALBA and modco derivatives BSCR 3 equivelant rating 
    portInput['Mapped_BSCR_Rating'] = np.where(
            (portInput['AIG Asset Class 3'] =='Derivative'),3, portInput['Mapped_BSCR_Rating'])
    # Private placement bonds NAIC rating + 2
    portInput['Mapped_BSCR_Rating'] = np.where(
            ((portInput['Mapped_BSCR_Rating'] == 0) & (portInput['NAIC Rating Band STAT 2'] != "SOURCE UNDEFINED") & ((portInput['Analytical Segment 3'] == 'High Grade Corps/BnkLns (Private)')|(portInput['Analytical Segment 3'] == 'High Yield Corps/BnkLns (Private)')|(portInput['Analytical Segment 3'] == 'Private High Yield')|(portInput['Analytical Segment 3'] == 'Private High Grade'))),
            portInput['NAIC Rating Band STAT 2'].str.replace('SOURCE UNDEFINED', '0').fillna(0). astype(int) + 2, portInput['Mapped_BSCR_Rating'])
    # no rating agency and no AIG derived rating -> 8
    portInput['Mapped_BSCR_Rating'] = np. where(
            ((portInput['Mapped_BSCR_Rating'] == 0) & (portInput[['S&P_num', 'Moody_num', 'Fitch_num', 'AIG_num']].sum(axis = 1) == 0)), 8, portInput['Mapped_BSCR_Rating'])     
    # no rating agency ratings, assign AIG derived rating
    portInput['Mapped_BSCR_Rating'] = np. where(
            ((portInput['Mapped_BSCR_Rating'] == 0) & ((portInput[['S&P_num', 'Moody_num', 'Fitch_num']].sum(axis = 1) == 0) & (portInput['AIG_num'] != 0))), portInput['AIG_num'], portInput['Mapped_BSCR_Rating'])
    # take max of rating agency ratings if they are available
    portInput['Mapped_BSCR_Rating'] = np. where(
            ((portInput['Mapped_BSCR_Rating'] == 0) & (portInput[['S&P_num', 'Moody_num', 'Fitch_num']].sum(axis = 1) != 0)), portInput[['S&P_num', 'Moody_num', 'Fitch_num']].max(axis = 1), portInput['Mapped_BSCR_Rating'])
     # NA RMBS CMBS   NAIC 1 rated 
    portInput['Mapped_BSCR_Rating'] = np. where(
            (((portInput['AIG Asset Class 3'] =='CMBS Agency') | (portInput['AIG Asset Class 3'] =='CMBS Non-Agency') | (portInput['AIG Asset Class 3'] =='RMBS Agency') | (portInput['AIG Asset Class 3'] =='RMBS Non-Agency')) & (portInput['NAIC Rating Band STAT 2'].str.replace('SOURCE UNDEFINED', '0').fillna(0). astype(int) ==1)\
             & (portInput['Mapped_BSCR_Rating'] >3)), 3, portInput['Mapped_BSCR_Rating'])
      
    portInput['Mapped_BSCR_Rating'] = np. where(
            (((portInput['AIG Asset Class 3'] =='CMBS Agency') | (portInput['AIG Asset Class 3'] =='CMBS Non-Agency') | (portInput['AIG Asset Class 3'] =='RMBS Agency') | (portInput['AIG Asset Class 3'] =='RMBS Non-Agency')) & (portInput['NAIC Rating Band STAT 2'].str.replace('SOURCE UNDEFINED', '0').fillna(0). astype(int) > 1)),\
              portInput['NAIC Rating Band STAT 2'].str.replace('SOURCE UNDEFINED', '0').fillna(0). astype(int) + 2, portInput['Mapped_BSCR_Rating'])      

  # Merge BMA asset class
    portInput = portInput.merge(AsMap, how='left', left_on=['AIG Asset Class 3'],
                                right_on=['AIG Asset class 3'])
    portInput['BMA Asset Category'] = np. where(
            portInput['Issuer Name'] =='LSTREET II, LLC', 'Alternatives', portInput['BMA Asset Category'])
    portInput['BMA Asset Category'] = np. where(
            ((portInput['BMA Asset Category'] =='Bonds Cash and Govt') & (portInput['AIG Asset Class 3'] !="Cash") & (portInput['Mapped_BSCR_Rating']>2)), 
            "Bonds", portInput['BMA Asset Category'])
    portInput['BMA Asset Category'] = np. where(
            ((portInput['AIG Asset Class 3'] =="Derivative")), 
            "Bonds", portInput['BMA Asset Category'])
    
  # Combine ratings and asset class
    portInput['Mapped_BSCR_Rating'] = portInput['Mapped_BSCR_Rating'].astype(int)
    portInput['BMA_Category'] = portInput['BMA Asset Category']
    portInput['BMA_Category'] = np.where(
            portInput['Issuer Name'] =='LSTREET II, LLC', 'ML III', portInput['BMA_Category'])
    portInput['BMA_Category'] = np.where(
            ((portInput['BMA_Category'] == 'Bonds') | (portInput['BMA_Category'] == 'CMBS') | (portInput['BMA_Category'] == 'RMBS')),
            portInput['BMA_Category'] + "_" + portInput['Mapped_BSCR_Rating'].map(str), portInput['BMA_Category'])
 
    portInput = portInput.rename(columns = {'BMA Asset Category': 'BMA_Asset_Class'})

   # Split out ML III Assets for concentration charge 
    portInput['Issuer Name'] = np.where(
            portInput['Issuer Name'] == 'LSTREET II, LLC', portInput['Issuer Name'] + '_' + portInput['Sec ID ID'].map(str), portInput['Issuer Name'])
    
   # Calculate cusip level charge
    portInput = portInput.merge(RcMap, how='left', left_on=['BMA_Category'],
                                right_on=['BMA_Category'])
    
    portInput['AssetCharge_Current'] = portInput['MV_USD_GAAP'] * portInput.Capital_factor_Current
    portInput['AssetCharge_Future'] = portInput['MV_USD_GAAP'] * portInput.Capital_factor_Future
    
    portInput['ConCharge_Current'] = portInput['MV_USD_GAAP'] * portInput.Concentration_factor_Current
    portInput['ConCharge_Future'] = portInput['MV_USD_GAAP'] * portInput.Concentration_factor_Future
  
    portInput['mv_dur']=portInput['Market Value USD GAAP'] * portInput['Effective Duration (WAMV)']
    
    # Existing Asset Charge    
    portInput['FIIndicator'] = portInput.BMA_Category.apply(
               lambda x: (0 if (x == 'Alternatives' or x == 'ML III' ) else 1))        
    portInput['EquityIndicator'] = portInput.BMA_Category.apply(
               lambda x: (1 if (x == 'Alternatives' or x == 'ML III') else 0))  
    
    portInput['FI Risk'] = 0
    portInput['FI Risk'] = np.where(
            portInput['FIIndicator'] == 1, portInput['AssetCharge_Current'], portInput['FI Risk'])

    
    portInput.fillna(0, inplace=True)

#   update ratings
    portInput['AIG Derived Rating Update'] = portInput['AIG Derived Rating']

    portInput['AIG Derived Rating Update'] = np.where(((portInput['AIG Asset Class 3'] == 'CMBS Agency') | (
            portInput['AIG Asset Class 3'] == 'CMBS Non-Agency') | (portInput['AIG Asset Class 3'] == 'RMBS Agency') | (
                                                               portInput['AIG Asset Class 3'] == 'RMBS Non-Agency')),
                                                      portInput['Derived Rating Modified'], portInput['AIG Derived Rating'])
    

####    Private Equity and ML III MV ajustment based on SPX
    portInput["pe_return"] = 1
#    for idx in range(0,len(portInput["Price Date"]),1):
#        if type(portInput["Price Date"][idx]) == 'str':
#            pe_base_date = datetime.datetime.strptime(portInput["Price Date"][idx],'%m/%d/%Y')
#        elif portInput["Price Date"][idx] == 0:
#            pe_base_date = valDate_base
#        else:
#            pe_base_date    = portInput["Price Date"][idx]
#        cusip_pe_return = IAL_App.eval_PE_return(eval_date, pe_base_date)
#        portInput["pe_return"] = np.where((portInput['AIG Asset Class 3']=="Private Equity Fund")&\
#                 (portInput.index == idx), 1 + cusip_pe_return , portInput["pe_return"])
    for idx in range(0,len(portInput["Price Date"]),1):
        if type(portInput["Price Date"][idx]) == 'str':
            pe_base_date = datetime.datetime.strptime(portInput["Price Date"][idx],'%m/%d/%Y')
        elif portInput["Price Date"][idx] == 0:
            pe_base_date = valDate_base
        else:
            pe_base_date    = portInput["Price Date"][idx]
        if portInput['AIG Asset Class 3'][idx]=="Private Equity Fund" :
            cusip_pe_return = IAL_App.eval_PE_return(eval_date, pe_base_date)
        else:
            cusip_pe_return = 0
        portInput["pe_return"] = np.where((portInput['AIG Asset Class 3']=="Private Equity Fund")&
                 (portInput.index == idx), 1 + cusip_pe_return , portInput["pe_return"])

        
    portInput['mv_adj']=portInput['Market Value USD GAAP']*portInput['pe_return']
#    portInput['mv_adj'] = np.where((portInput['AIG Asset Class 3']=="ML-III B-Notes"),portInput['Market Value USD GAAP'] * (1 + pe_return),\
#             portInput['Market Value USD GAAP'])
    
    portInput['Eq Risk_Current'] = 0
    portInput['Eq Risk_Current'] = np.where(
            portInput['EquityIndicator'] == 1, portInput['mv_adj']*portInput.Capital_factor_Current, portInput['Eq Risk_Current'])
    portInput['AssetCharge_Current'] = np.where(portInput['EquityIndicator'] == 1,portInput['Eq Risk_Current'],portInput['AssetCharge_Current'])

    portInput['Eq Risk_Future'] = 0
    portInput['Eq Risk_Future'] = np.where(
            portInput['EquityIndicator'] == 1, portInput['mv_adj']*portInput.Capital_factor_Future, portInput['Eq Risk_Future'])
    portInput['AssetCharge_Future'] = np.where(portInput['EquityIndicator'] == 1,portInput['Eq Risk_Future'],portInput['AssetCharge_Future'])

  
    
#   MV * Duration Calculations
    portInput['mv_dur']         = portInput['mv_adj'] * portInput['Effective Duration (WAMV)'] 
    portInput['acc_int']        = portInput['Market Value with Accrued Int USD GAAP'] - portInput['Market Value USD GAAP']
    portInput['mv_acc_int_adj'] = portInput['mv_adj'] + portInput['acc_int']

    ### === Prepare Asset Summary File === ###
    
    # temporarily write to file for validation purpose
    # portWriter = pd.ExcelWriter('./output/FRL_0228.xlsx')
    # portInput.to_excel(portWriter, sheet_name='asset_port', index=False)
    # portWriter.save()

    groupByFields = ['Category', 'AIG Asset Class 3', 'AIG Derived Rating Update']

    reportFields = {"YTW": "Market Value USD GAAP", \
                    "OAS": "Market Value USD GAAP", \
                    "Effective Duration (WAMV)": "Market Value USD GAAP", \
                    "Effective Convexity": "Market Value USD GAAP", \
                    "Spread Duration": "Market Value USD GAAP", \
                    "Spread Convexity": "Market Value USD GAAP", \
                    "WAL": "Quantity", \
                    "KRD 1M": "Market Value USD GAAP", \
                    "KRD 2M": "Market Value USD GAAP", \
                    "KRD 3M": "Market Value USD GAAP", \
                    "KRD 6M": "Market Value USD GAAP", \
                    "KRD 9M": "Market Value USD GAAP", \
                    "KRD 1Y": "Market Value USD GAAP", \
                    "KRD 2Y": "Market Value USD GAAP", \
                    "KRD 3Y": "Market Value USD GAAP", \
                    "KRD 5Y": "Market Value USD GAAP", \
                    "KRD 7Y": "Market Value USD GAAP", \
                    "KRD 10Y": "Market Value USD GAAP", \
                    "KRD 15Y": "Market Value USD GAAP", \
                    "KRD 20Y": "Market Value USD GAAP", \
                    "KRD 30Y": "Market Value USD GAAP"}

    sumFields = groupByFields.copy()
    sumFields.append('Market Value USD GAAP')
    mv = portInput.loc[pd.notnull(portInput['Market Value USD GAAP']), sumFields].groupby(groupByFields).sum()

    sumFields = groupByFields.copy()
    sumFields.append('Market Value with Accrued Int USD GAAP')
    mv_acc = portInput.loc[pd.notnull(portInput['Market Value with Accrued Int USD GAAP']), sumFields].groupby(
        groupByFields).sum()

    sumFields = groupByFields.copy()
    sumFields.append('Book Value USD STAT')
    bv = portInput.loc[pd.notnull(portInput['Book Value USD STAT']), sumFields].groupby(groupByFields).sum()

    sumFields = groupByFields.copy()
    sumFields.append('Book Value With Accrued Int USD STAT')
    bv_acc = portInput.loc[pd.notnull(portInput['Book Value With Accrued Int USD STAT']), sumFields].groupby(
        groupByFields).sum()

    merge = pd.merge(left=mv, left_index=True, right=mv_acc, right_index=True, how='outer')
    merge = pd.merge(left=merge, left_index=True, right=bv, right_index=True, how='outer')
    merge = pd.merge(left=merge, left_index=True, right=bv_acc, right_index=True, how='outer')

    for key, val in reportFields.items():
        ds = portInput[pd.notnull(portInput[key])].groupby(groupByFields).apply(wavg(key, val))
        df = pd.DataFrame(data=ds, columns=[key])
        merge = pd.merge(left=merge, left_index=True, right=df, right_index=True, how='outer')

    asset_count = merge.shape[0]
    cal_category    = []
    cal_asset_class = []
    cal_rating      = []
    
    merge.fillna(0, inplace=True)
    
    for idx in range(0, asset_count, 1):
        cal_category.append(merge.index[idx][0])
        cal_asset_class.append(merge.index[idx][1])
        cal_rating.append(merge.index[idx][2])
    
    merge['category_f']     = cal_category
    merge['asset_class_f'] = cal_asset_class
    merge['rating_f']       = cal_rating


#    Private Equity MV ajustment based on SPX
    sumFields = groupByFields.copy()
    sumFields.append('mv_adj')
    mv_adj = portInput.loc[pd.notnull(portInput['mv_adj']), sumFields].groupby(
        groupByFields).sum()
    merge = pd.merge(left=merge, left_index=True, right=mv_adj, right_index=True, how='outer')
#    pe_return = IAL_App.eval_PE_return(eval_date, valDate_base)
#    mv_adj = []
#    mv_acc = []
#    for index, row in merge.iterrows():
#        if row['asset_class_f'] == "Private Equity Fund" or row['asset_class_f'] == "ML-III B-Notes":
#            mv_adj.append(row['Market Value USD GAAP'] * (1 + pe_return) )
#            mv_acc
#        else:
#            mv_adj.append(row['Market Value USD GAAP'])

#    merge['mv_adj'] = mv_adj
    
#   MV * Duration Calculations
    merge['mv_dur']         = merge['mv_adj'] * merge['Effective Duration (WAMV)'] 
    merge['acc_int']        = merge['Market Value with Accrued Int USD GAAP'] - merge['Market Value USD GAAP']
    merge['mv_acc_int_adj'] = merge['mv_adj'] + merge['acc_int']

    
#   Remove non ALBA derivative
    for i in range(0, len(merge.loc[('ModCo','Derivative'), "Market Value USD GAAP"]) , 1):
        merge.loc[('ModCo','Derivative'), "Market Value USD GAAP"][i]                  = 0
        merge.loc[('ModCo','Derivative'), "Market Value with Accrued Int USD GAAP"][i] = 0
        merge.loc[('ModCo','Derivative'), "mv_adj"][i]         = ModCo_Non_IR_Der   #to get derivative value
        merge.loc[('ModCo','Derivative'), "mv_acc_int_adj"][i] = 0
        
    for i in range(0, len(merge.loc[('LPT','Derivative'), "Market Value USD GAAP"]) , 1):
        merge.loc[('LPT','Derivative'), "Market Value USD GAAP"][i]                  = 0
        merge.loc[('LPT','Derivative'), "Market Value with Accrued Int USD GAAP"][i] = 0
        merge.loc[('LPT','Derivative'), "mv_adj"][i]         = LPT_Non_IR_Der   #to get derivative value
        merge.loc[('LPT','Derivative'), "mv_acc_int_adj"][i] = 0
        
#   Rectify the derivative on 1-day lag  
    try:
        asset_file_T_plus_1 = pd.ExcelFile(asset_fileName_T_plus_1)
        asset_T_plus_1 = pd.read_excel(asset_file_T_plus_1, sheet_name='AssetSummaryFromPython')
        
        der_T_plus_1_mv_USD_GAAP    = asset_T_plus_1.loc[(asset_T_plus_1.Category == 'ModCo') & (asset_T_plus_1['AIG Asset Class 3'] == 'Derivative'), "Market Value USD GAAP"].sum()
        der_T_plus_1_mv_adj         = asset_T_plus_1.loc[(asset_T_plus_1.Category == 'ModCo') & (asset_T_plus_1['AIG Asset Class 3'] == 'Derivative'), "mv_adj"].sum()
        der_T_plus_1_mv_acc_int_adj = asset_T_plus_1.loc[(asset_T_plus_1.Category == 'ModCo') & (asset_T_plus_1['AIG Asset Class 3'] == 'Derivative'), "mv_acc_int_adj"].sum()
                
        ori_der = merge.loc[('ModCo','Derivative'), "Market Value USD GAAP"].sum()
        print(str(eval_date) + '_Original: ' + str(ori_der) )
        print(str(eval_date) + '_Revised: '  + str(der_T_plus_1_mv_adj) )
        
        for i in range(0, len(merge.loc[('ModCo','Derivative'), "Market Value USD GAAP"]) , 1):            
            if i == 0:
                merge.loc[('ModCo','Derivative'), "Market Value USD GAAP"][i] = der_T_plus_1_mv_USD_GAAP
                merge.loc[('ModCo','Derivative'), "mv_adj"][i]                = der_T_plus_1_mv_adj
                merge.loc[('ModCo','Derivative'), "mv_acc_int_adj"][i]        = der_T_plus_1_mv_acc_int_adj
            else:   # if there is more than one row of ModCo derivatives
                merge.loc[('ModCo','Derivative'), "Market Value USD GAAP"][i] = 0
                merge.loc[('ModCo','Derivative'), "mv_adj"][i]                = 0
                merge.loc[('ModCo','Derivative'), "mv_acc_int_adj"][i]        = 0
                
        if output == 1:
            out_file = fileName[0:25] + "_summary_Der_revised.xlsx" 
            assetSummary = pd.ExcelWriter(out_file)
            merge.to_excel(assetSummary, sheet_name='AssetSummaryFromPython', index=True, merge_cells=False)
            assetSummary.save()
            """
            wb=xlw.Book.caller()
            wb.sheets[0].range('A1').value = merge
            wb.sheets[1].range('A1').value = portInput
            """
        return merge
         
    except:              
        if output == 1:
            out_file = fileName[0:25] + "_summary.xlsx"
            assetSummary = pd.ExcelWriter(out_file)
            merge.to_excel(assetSummary, sheet_name='AssetSummaryFromPython', index=True, merge_cells=False)
            assetSummary.save()
            """
            wb=xlw.Book.caller()
            wb.sheets[0].range('A1').value = merge
            wb.sheets[1].range('A1').value = portInput
            """
#        return merge
#        sumFields = groupByFields.copy()
#        sumFields.append('FI Risk')
#        FI_rc = portInput.loc[pd.notnull(portInput['FI Risk']), sumFields].groupby(groupByFields).sum()
#       
#        sumFields = groupByFields.copy()
#        sumFields.append('Eq Risk_Current')
#        Eq_rc_current = portInput.loc[pd.notnull(portInput['Eq Risk_Current']), sumFields].groupby(groupByFields).sum()
#
#        sumFields = groupByFields.copy()
#        sumFields.append('Eq Risk_Future')
#        Eq_rc_future  = portInput.loc[pd.notnull(portInput['Eq Risk_Future']), sumFields].groupby(groupByFields).sum()
#
#        
#        merge = pd.merge(left=merge, left_index=True, right=FI_rc, right_index=True, how='outer')
#        merge = pd.merge(left=merge, left_index=True, right=Eq_rc_current, right_index=True, how='outer')
#        merge = pd.merge(left=merge, left_index=True, right=Eq_rc_future, right_index=True, how='outer')
        
        return portInput

#def actual_portfolio_feed(workDir, fileName, AssetRiskCharge):
    
#    os.chdir(workDir)
    
#    portFile = pd.ExcelFile(fileName)
#    portInput = pd.read_excel(portFile, sheet_name='Microstrategy_Holdings')
               
#    portInput = pd.merge(portInput, AssetRiskCharge, how ='left', left_on='BMA_Category', right_on='BMA_Category')
   
#    portInput['AssetCharge_Current'] = portInput.MV_USD_GAAP * portInput.Capital_factor_Current
#    portInput['AssetCharge_Future'] = portInput.MV_USD_GAAP * portInput.Capital_factor_Future
    
#    portInput['mv_dur']=portInput.MV_USD_GAAP * portInput['Effective Duration (WAMV)']
    
    # Existing Asset Charge    
#    portInput['FIIndicator'] = portInput.BMA_Category.apply(
#               lambda x: (0 if (x == 'Alternatives' or x == 'ML III' ) else 1))        
#    portInput['EquityIndicator'] = portInput.BMA_Category.apply(
#               lambda x: (1 if (x == 'Alternatives' or x == 'ML III') else 0))    
         
#    return portInput
    
def Set_weighted_average_OAS(valDate, EBS_Cal_Dates_all, asset_workDir):    
    
    os.chdir(asset_workDir)
    colNames = ['ValDate','OAS_ModCo','OAS_LPT']
    eval_date = [valDate] + EBS_Cal_Dates_all
    credit_spread = pd.DataFrame([],columns = colNames)
    
    for val_date in eval_date:
        if val_date < datetime.datetime(2019,12,31) :
            asset_fileName = r'.\Asset_Holdings_' + val_date.strftime('%Y%m%d') + '_update_oas.xlsx'
            portFile       = pd.ExcelFile(asset_fileName)
            portInput = pd.read_excel(portFile, sheet_name='AssetSummaryFromPython')
            portInput      = portInput.dropna(axis=0, how='all')
            portInput      = portInput.dropna(axis=1, how='all')
            portInput.sort_values('Market Value USD GAAP',inplace=True)
            portInput.drop_duplicates('Lot Number Unique ID',inplace=True)
            portInput.fillna(value=0,inplace=True)
            portInput['oas_mv']=portInput['OAS']*portInput['Market Value USD GAAP']
            asset_sort_LR         = portInput[(portInput['AIG Asset Class 1'] == "Fixed Income")\
                                           &((portInput['Owning Entity Name'] == "American General Life Insurance Company")\
                                           |(portInput['Owning Entity Name'] == "The United States Life Insurance Company in the City of New York")\
                                           |(portInput['Owning Entity Name'] == "The Variable Annuity Life Insurance Company")\
                                           )]
            asset_sort_PC         = portInput[(portInput['AIG Asset Class 1'] == "Fixed Income")\
                                           &((portInput['Owning Entity Name'] == "National Union Fire Insurance Company of Pittsburgh, Pa.")\
                                           |(portInput['Owning Entity Name'] == "Lexington Insurance Company")\
                                           |(portInput['Owning Entity Name'] == "American Home Assurance Company")\
                                           )]

            weighted_OAS_LR  = asset_sort_LR['oas_mv'].sum()/asset_sort_LR['Market Value USD GAAP'].sum() 
            weighted_OAS_PC  = asset_sort_PC['oas_mv'].sum()/asset_sort_PC['Market Value USD GAAP'].sum() 

            each_credit_spread = [val_date, weighted_OAS_LR, weighted_OAS_PC]
        else:## Back-solved OAS from 12/31/2019
            oas_file = r'.\OAS-20Q1.xlsx'
            portFile = pd.ExcelFile(oas_file)
            portInput = pd.read_excel(portFile, sheet_name ='Sheet1')
            weighted_OAS = portInput[portInput['dates']<=val_date]['OAS'].sum()
            if val_date >= datetime.datetime(2020,3,31):
                ModCo_OAS   = weighted_OAS + portInput[portInput['dates']<=val_date]['OAS_ModCo'].sum()
                LPT_OAS     = weighted_OAS + portInput[portInput['dates']<=val_date]['OAS_LPT'].sum()
                print("OAS_ModCo: "+str(val_date)+" "+str(ModCo_OAS) )
                print("OAS_LPT: "+str(val_date)+" "+str(LPT_OAS) )            
                each_credit_spread = [val_date, ModCo_OAS,LPT_OAS]
            else:   
                print("OAS: "+str(val_date)+" "+str(weighted_OAS) )
                each_credit_spread = [val_date, weighted_OAS, weighted_OAS]
        credit_spread = credit_spread.append(pd.DataFrame([each_credit_spread], columns = colNames), ignore_index = True)
    return credit_spread

def load_asset_OAS(EBS_Cal_Dates_all, asset_workDir, asset_attribution):
    os.chdir(asset_workDir)
    dataframe= pd.read_excel("L:\\DSA Re\\Workspace\\Production\\EBS Dashboard\\Python_Code\\Asset_Holding_Feed\\OAS-20Q1.xlsx")
    date_col = dataframe['dates']
    dates    = EBS_Cal_Dates_all[1:]
    for each_date in dates:
        att_data = asset_attribution[each_date]
        for i in range(0,5,1):
            if att_data[i]["Group"] == "ModCo" :
                ModCo_credit_01    = att_data[i]["x credit_01"]
                ModCo_oas_att      = att_data[i]["OAS attribution"]
            elif att_data[i]["Group"] == "LPT" :
                LPT_credit_01      = att_data[i]["x credit_01"]
                LPT_oas_att        = att_data[i]["OAS attribution"]            
        ModCo_oas_change = -ModCo_oas_att/ModCo_credit_01*10000
        LPT_oas_change   = -LPT_oas_att/LPT_credit_01*10000
        if each_date in list(date_col):##Update existing OAS
            dataframe['OAS_ModCo'] = np.where((dataframe['dates']==each_date),ModCo_oas_change,dataframe['OAS_ModCo'])
            dataframe['OAS_LPT'] = np.where((dataframe['dates']==each_date),LPT_oas_change,dataframe['OAS_LPT'])
            dataframe.to_excel("L:\\DSA Re\\Workspace\\Production\\EBS Dashboard\\Python_Code\\Asset_Holding_Feed\\OAS-20Q1.xlsx",index = False)
        else:##store OAS for new dates
            oasbook  = load_workbook(filename="OAS-20Q1.xlsx")
            sheet    = oasbook.get_sheet_by_name('Sheet1')
            a = sheet.max_row
            print(each_date)
            print(a)
            sheet.cell(row=a+1,column=1).value=each_date
            sheet.cell(row=a+1,column=3).value=ModCo_oas_change
            sheet.cell(row=a+1,column=4).value=LPT_oas_change 
            print(sheet.cell(row=a+1,column=1).value)                 
            oasbook.save("OAS-20Q1.xlsx")
            oasbook.close()
    return
        
def update_illiquid_oas(EBS_Cal_Dates_all, asset_workDir, market_factor,Price_Date, write_to_excel = 1):
     month_end_date = 0
     
     for each_date in EBS_Cal_Dates_all:
        print('Updating illiquid assets OAS for ' + str(each_date))

        if each_date >= datetime.datetime(2020,6,1):
            price_date_1 = datetime.datetime(2020,5,29)
        else:
            price_date_1 = each_date + MonthEnd(-1) 
        
        os.chdir(asset_workDir)
        asset_fileName = r'.\Asset_Holdings_' + each_date.strftime('%Y%m%d') + '.xlsx'
        portFile = pd.ExcelFile(asset_fileName)
        portInput = pd.read_excel(portFile, sheet_name='DSA RE Holdings', skiprows=[0, 1, 2, 3, 4, 5, 6]) 
         
        if each_date in list(Price_Date):
            pass
        elif each_date == datetime.datetime(2020,2,28):
            pass
        else:
            if price_date_1 != month_end_date:
                print('Loading price date illiquid assets OAS as of ' + str(price_date_1))
                asset_fileName_price_date = r'.\Asset_Holdings_' + price_date_1.strftime('%Y%m%d') + '.xlsx'
                portFile_price_date = pd.ExcelFile(asset_fileName_price_date)
                portInput_price_date = pd.read_excel(portFile_price_date, sheet_name='DSA RE Holdings', skiprows=[0, 1, 2, 3, 4, 5, 6]) 
                portInput_price_date = portInput_price_date[['Lot Number Unique ID','OAS']].rename(columns = {'OAS': 'OAS_price_date'})
                
            portInput = portInput.merge(portInput_price_date, how = 'left', left_on = 'Lot Number Unique ID', right_on = 'Lot Number Unique ID')
                             
            OAS_Change_1 = 0.5 * (market_factor[market_factor['val_date'] == each_date]['Credit_Spread'].values[0] - market_factor[market_factor['val_date'] == price_date_1]['Credit_Spread'].values[0])
            print(OAS_Change_1)
            
    ## remove condition of wamv not null
            portInput['OAS'] = np.where((portInput['Price Date'] <= price_date_1) & (portInput['AIG Asset Class 1'] == 'Fixed Income')  & (portInput['Effective Duration (WAMV)'] != 0) & (portInput['Issuer Name'] != 'LSTREET II, LLC') & (portInput['OAS_price_date'].notnull()),
                                        portInput['OAS_price_date'] + OAS_Change_1,
                                        portInput['OAS'] )

            # Illiquidity impact estimation
            price_date_0 = each_date + MonthEnd(-2)
            Yield_Change_0 = market_factor[market_factor['val_date'] == each_date]['IR'].values[0] - market_factor[market_factor['val_date'] == price_date_0]['IR'].values[0] + 0.5 * (market_factor[market_factor['val_date'] == each_date]['Credit_Spread'].values[0]/10000 - market_factor[market_factor['val_date'] == price_date_0]['Credit_Spread'].values[0]/10000)
                           
            Yield_Change_1 = market_factor[market_factor['val_date'] == each_date]['IR'].values[0] - market_factor[market_factor['val_date'] == price_date_1]['IR'].values[0] + 0.5 * (market_factor[market_factor['val_date'] == each_date]['Credit_Spread'].values[0]/10000 - market_factor[market_factor['val_date'] == price_date_1]['Credit_Spread'].values[0]/10000)
    
            Initial_mv_acc_int = portInput.loc[(portInput['Price Date'] <= price_date_1) & (portInput['AIG Asset Class 1'] == 'Fixed Income') & (portInput['Effective Duration (WAMV)'].notnull() ) & (portInput['Effective Duration (WAMV)'] != 0), 'Market Value with Accrued Int USD GAAP'].sum()
                
            portInput.loc[(portInput['Price Date'] > price_date_0) & (portInput['Price Date'] <= price_date_1) & (portInput['AIG Asset Class 1'] == 'Fixed Income') & (portInput['Effective Duration (WAMV)'].notnull() ) & (portInput['Effective Duration (WAMV)'] != 0) & (portInput['AIG Asset Class 3'] != 'ML-III B-Notes'), "Market Value with Accrued Int USD GAAP"] = portInput['Market Value with Accrued Int USD GAAP'] * (1 + portInput['Effective Duration (WAMV)'] * -Yield_Change_1) 
            portInput.loc[(portInput['Price Date'] <= price_date_0) & (portInput['AIG Asset Class 1'] == 'Fixed Income') & (portInput['Effective Duration (WAMV)'].notnull() ) & (portInput['Effective Duration (WAMV)'] != 0)& (portInput['AIG Asset Class 3'] != 'ML-III B-Notes'), "Market Value with Accrued Int USD GAAP"] = portInput['Market Value with Accrued Int USD GAAP'] * (1 + portInput['Effective Duration (WAMV)'] * -Yield_Change_0) 
                        
            total_illiquid_adj = portInput.loc[(portInput['Price Date'] <= price_date_1) & (portInput['AIG Asset Class 1'] == 'Fixed Income') & (portInput['Effective Duration (WAMV)'].notnull() ) & (portInput['Effective Duration (WAMV)'] != 0), 'Market Value with Accrued Int USD GAAP'].sum() - Initial_mv_acc_int
           
            portInput.fillna(0, inplace=True)
            portInput['Market Value USD GAAP'] = portInput['Market Value with Accrued Int USD GAAP'] - portInput['Accrued Int USD GAAP']
            
            print(each_date.strftime('%Y%m%d') + 'Yield_Change_' + price_date_0.strftime('%Y%m%d') + ': ' + str(Yield_Change_0))
            print(each_date.strftime('%Y%m%d') + 'Yield_Change_' + price_date_1.strftime('%Y%m%d') + ': ' + str(Yield_Change_1))
            print(each_date.strftime('%Y%m%d') + '_Initial_mv_acc_int: ' + str(Initial_mv_acc_int))
            print(each_date.strftime('%Y%m%d') + '_total_illiquid_adj: ' + str(total_illiquid_adj))
                                

        out_file = asset_fileName[0:25] + "_update_oas_old_format.xlsx"
        assetSummary = pd.ExcelWriter(out_file)
        portInput.to_excel(assetSummary, sheet_name='AssetSummaryFromPython', index=True, merge_cells=False)
        assetSummary.save()
        
        month_end_date = price_date_1


        
def Asset_Adjustment_feed(AssetAdjustment):
    
    AssetRiskCharge = pd.DataFrame(BSCR_Cofig.BSCR_Asset_Risk_Charge_v1).transpose()
    AssetRiskCharge['BMA_Category'] = AssetRiskCharge.index 
    
    AssetAdjustment = pd.merge(AssetAdjustment, AssetRiskCharge, how ='left', left_on='BMA_Category', right_on='BMA_Category')
    
    AssetAdjustment['AssetCharge_Current'] = AssetAdjustment.MV_USD_GAAP * AssetAdjustment.Risk_Charge

    AssetAdjustment['AssetCharge_Future'] = AssetAdjustment.MV_USD_GAAP * AssetAdjustment.Risk_Charge
    
    # Existing Asset Charge    
    AssetAdjustment['FIIndicator'] = AssetAdjustment.BMA_Category.apply(
               lambda x: (0 if x == 'LOC' or x == 'Alternatives' else 1))        
    AssetAdjustment['EquityIndicator'] = AssetAdjustment.BMA_Category.apply(
               lambda x: (1 if x == 'LOC' or x == 'Alternatives' else 0))  
    
    return AssetAdjustment
    
def actual_portfolio_feed(eval_date, valDate_base, workDir, fileName, ALBA_fileName, output):
    
    def wavg(val_col_name, wt_col_name):
        def inner(group):
            return (group[val_col_name] * group[wt_col_name]).sum() / group[wt_col_name].sum()

        inner.__name__ = 'wtd_avg'
        return inner
    
    curr_dir = os.getcwd()
    os.chdir(workDir)

    portFile = pd.ExcelFile(fileName)
    #mapFile = pd.ExcelFile(mapping)
    
    try:
        albaFile = pd.ExcelFile(ALBA_fileName)
        ALBA = pd.read_excel(albaFile)
    except:
        pass
    
    # if estimate, sheet_name='DSA RE Holdings'
    portInput = pd.read_excel(portFile, sheet_name='Microstrategy_Holdings', skiprows=[0, 1, 2, 3, 4, 5, 6])  
    leMap = pd.DataFrame(Rating_Cofig.Legal_Entity)

    RaMap_sp = pd.DataFrame(Rating_Cofig.SP_Rating, index=['BSCR rating_SP']).transpose()
    RaMap_sp["S&P"] = RaMap_sp.index
    
    RaMap_moodys = pd.DataFrame(Rating_Cofig.Moodys_Rating, index=['BSCR rating_Moodys']).transpose()
    RaMap_moodys["Moody's"] = RaMap_moodys.index
    
    RaMap_fitch = pd.DataFrame(Rating_Cofig.Fitch_Rating, index=['BSCR rating_Fitch']).transpose()
    RaMap_fitch["Fitch"] = RaMap_fitch.index
    
    RaMap_aig = pd.DataFrame(Rating_Cofig.AIG_Rating, index=['BSCR rating_AIG']).transpose()
    RaMap_aig["AIG Rating"] = RaMap_aig.index
    
    AsMap = pd.DataFrame(Rating_Cofig.AC_Mapping_to_BMA, index=['BMA Asset Category']).transpose()
    AsMap['AIG Asset class 3'] = AsMap.index 
    
    RcMap = pd.DataFrame(BSCR_Cofig.BSCR_Asset_Risk_Charge_v1).transpose()
    RcMap['BMA_Category'] = RcMap.index
    
    portInput = portInput.dropna(axis=0, how='all')
    portInput = portInput.dropna(axis=1, how='all')

    portInput = portInput.merge(leMap, how='left', left_on=['Owning Entity Name', 'Owning Entity GEMS ID'],
                                right_on=['Owning Entity Name', 'Owning Entity GEMS ID'])
    
    ratingMap = pd.DataFrame(Rating_Cofig.Rating)

    portInput = portInput.merge(ratingMap, how='left', left_on=['NAIC Rating', 'AIG Derived Rating'],
                                right_on=['NAIC Rating', 'AIG Derived Rating'])
    
    # Split out ML-III B Notes
    portInput['AIG Asset Class 3'] = np.where((portInput['Issuer Name'] == 'LSTREET II, LLC'), "ML-III B-Notes",
                                              portInput['AIG Asset Class 3'])

    # Handling surplus account
    portInput['SurplusAccount'] = 'Not-Surplus'

    portInput['SurplusAccount'] = np.where(
        ((portInput['Category'] == "Surplus") & (portInput['Strategy (Hld) Long Desc'] == "SAM RE OVERSEAS POOL LIFE")),
        "Long Term Surplus", portInput['SurplusAccount'])

    portInput['SurplusAccount'] = np.where(
        ((portInput['Category'] == "Surplus") & (portInput['Strategy (Hld) Long Desc'] == "SAM RE OVERSEAS POOL PC")),
        "General Surplus", portInput['SurplusAccount'])

    portInput['SurplusAccount'] = np.where(
        ((portInput['SurplusAccount'] == 'Not-Surplus') &
         (portInput['Category'] == "Surplus") & (portInput['Encumbrance Program Level 4 Desc'] == 'DSA REINSURANCE PC')),
        "General Surplus", portInput['SurplusAccount'])

    portInput['SurplusAccount'] = np.where(
        ((portInput['SurplusAccount'] == 'Not-Surplus') &
         (portInput['Category'] == "Surplus") & (portInput['Encumbrance Program Level 4 Desc'] == 'DSA REINSURANCE LR')),
        "Long Term Surplus", portInput['SurplusAccount'])

    portInput['SurplusAccount'] = np.where(
        ((portInput['SurplusAccount'] == 'Not-Surplus') &
         (portInput['Category'] == "Surplus") & (
                 portInput['Encumbrance Program Level 4 Desc'] == 'DSA RE/AGL Trust- DSA REINSURANCE LR')),
        "Long Term Surplus", portInput['SurplusAccount'])

    portInput['SurplusAccount'] = np.where(
        ((portInput['SurplusAccount'] == 'Not-Surplus') &
         (portInput['Category'] == "Surplus") & (
                 portInput['Encumbrance Program Level 4 Desc'] == 'DSA RE/USL Trust - DSA REINSURANCE LR')),
        "Long Term Surplus", portInput['SurplusAccount'])

    portInput['SurplusAccount'] = np.where(
        ((portInput['SurplusAccount'] == 'Not-Surplus') &
         (portInput['Category'] == "Surplus") & (
                 portInput['Encumbrance Program Level 4 Desc'] == 'DSA RE/VALIC Trust - DSA REINSURANCE LR')),
        "Long Term Surplus", portInput['SurplusAccount'])

    portInput['Category'] = np.where(
        ((portInput['SurplusAccount'] == 'Not-Surplus') &
         (portInput['Category'] == "Surplus") & (portInput['Encumbrance Program Level 4 Desc'] == 'SOURCE UNDEFINED')),
        "Long Term Surplus", portInput['Category'])

    portInput.dropna(axis=0, how='all', subset=['Category'], inplace=True)

    portInput['Category'] = np.where((portInput['SurplusAccount'] != 'Not-Surplus'),
                                     portInput['SurplusAccount'], portInput['Category'])
         
    portInput['Fort Re Corp Segment'] = portInput['Category']

    # clean up fake booking due to accounting system migration
    portInput = portInput[[not v.startswith('F-') for (i, v) in portInput['Portfolio (Source) Long Name'].iteritems()]]
    
    if valDate_base == datetime.datetime(2018, 12, 31, 0, 0):
        # zero out non-ALBA derivatives
        portInput['Market Value USD GAAP'] = np.where(
                ((portInput['AIG Asset Class 3'] =='Derivative') & (portInput['Owning Entity Name'] !='American International Reinsurance Company, Ltd.')),
                0, portInput['Market Value USD GAAP'])
        
        ###-----Only for 4Q18 72M ModCo to LT surplus reclass-----###
        portInput['Category'] = np.where(
        ((portInput['Category'] == 'ModCo') &
         (portInput['Base Line Of Business Code'] == 'SOURCE UNDEFINED')),
        'Long Term Surplus', portInput['Category'])
    
    portInput['MV_USD_GAAP'] = portInput['Market Value USD GAAP']
    
    # Assign number scale to ratings
    portInput = portInput.merge(RaMap_sp, how = 'left', left_on = 'S&P Rating', right_on = 'S&P')\
                        .rename(columns = {'BSCR rating_SP': 'S&P_num'})\
                        .drop(columns = 'S&P')
 
    portInput = portInput.merge(RaMap_moodys, how = 'left', left_on = "Moody's Rating", right_on = "Moody's")\
                        .rename(columns = {'BSCR rating_Moodys': 'Moody_num'})\
                        .drop(columns = "Moody's")
                        
    portInput = portInput.merge(RaMap_fitch, how = 'left', left_on = 'Fitch Rating', right_on = 'Fitch')\
                        .rename(columns = {'BSCR rating_Fitch': 'Fitch_num'})\
                        .drop(columns = 'Fitch')

    portInput = portInput.merge(RaMap_aig, how = 'left', left_on = 'AIG Derived Rating', right_on = 'AIG Rating')\
                        .rename(columns = {'BSCR rating_AIG': 'AIG_num'})\
                        .drop(columns = 'AIG Rating')

    # Assign BSCR rating
    portInput['Mapped_BSCR_Rating'] = 0
   # ALBA and modco derivatives BSCR 3 equivelant rating 
    portInput['Mapped_BSCR_Rating'] = np.where(
            (portInput['AIG Asset Class 3'] =='Derivative'),3, portInput['Mapped_BSCR_Rating'])
    # Private placement bonds NAIC rating + 2
    portInput['Mapped_BSCR_Rating'] = np.where(
            ((portInput['Mapped_BSCR_Rating'] == 0) & (portInput['NAIC Rating Band STAT 2'] != "SOURCE UNDEFINED") & ((portInput['Analytical Segment 3'] == 'High Grade Corps/BnkLns (Private)')|(portInput['Analytical Segment 3'] == 'High Yield Corps/BnkLns (Private)')|(portInput['Analytical Segment 3'] == 'Private High Yield')|(portInput['Analytical Segment 3'] == 'Private High Grade'))),
            portInput['NAIC Rating Band STAT 2'].str.replace('SOURCE UNDEFINED', '0').fillna(0). astype(int) + 2, portInput['Mapped_BSCR_Rating'])
    # no rating agency and no AIG derived rating -> 8
    portInput['Mapped_BSCR_Rating'] = np. where(
            ((portInput['Mapped_BSCR_Rating'] == 0) & (portInput[['S&P_num', 'Moody_num', 'Fitch_num', 'AIG_num']].sum(axis = 1) == 0)), 8, portInput['Mapped_BSCR_Rating'])     
    # no rating agency ratings, assign AIG derived rating
    portInput['Mapped_BSCR_Rating'] = np. where(
            ((portInput['Mapped_BSCR_Rating'] == 0) & ((portInput[['S&P_num', 'Moody_num', 'Fitch_num']].sum(axis = 1) == 0) & (portInput['AIG_num'] != 0))), portInput['AIG_num'], portInput['Mapped_BSCR_Rating'])
    # take max of rating agency ratings if they are available
    portInput['Mapped_BSCR_Rating'] = np. where(
            ((portInput['Mapped_BSCR_Rating'] == 0) & (portInput[['S&P_num', 'Moody_num', 'Fitch_num']].sum(axis = 1) != 0)), portInput[['S&P_num', 'Moody_num', 'Fitch_num']].max(axis = 1), portInput['Mapped_BSCR_Rating'])
     # NA RMBS CMBS   NAIC 1 rated 
    portInput['Mapped_BSCR_Rating'] = np. where(
            (((portInput['AIG Asset Class 3'] =='CMBS Agency') | (portInput['AIG Asset Class 3'] =='CMBS Non-Agency') | (portInput['AIG Asset Class 3'] =='RMBS Agency') | (portInput['AIG Asset Class 3'] =='RMBS Non-Agency')) & (portInput['NAIC Rating Band STAT 2'].str.replace('SOURCE UNDEFINED', '0').fillna(0). astype(int) ==1)\
             & (portInput['Mapped_BSCR_Rating'] >3)), 3, portInput['Mapped_BSCR_Rating'])
      
    portInput['Mapped_BSCR_Rating'] = np. where(
            (((portInput['AIG Asset Class 3'] =='CMBS Agency') | (portInput['AIG Asset Class 3'] =='CMBS Non-Agency') | (portInput['AIG Asset Class 3'] =='RMBS Agency') | (portInput['AIG Asset Class 3'] =='RMBS Non-Agency')) & (portInput['NAIC Rating Band STAT 2'].str.replace('SOURCE UNDEFINED', '0').fillna(0). astype(int) > 1)),\
              portInput['NAIC Rating Band STAT 2'].str.replace('SOURCE UNDEFINED', '0').fillna(0). astype(int) + 2, portInput['Mapped_BSCR_Rating'])      

  # Merge BMA asset class
    portInput = portInput.merge(AsMap, how='left', left_on=['AIG Asset Class 3'],
                                right_on=['AIG Asset class 3']).drop(columns = 'AIG Asset class 3')
    portInput['BMA Asset Category'] = np. where(
            portInput['Issuer Name'] =='LSTREET II, LLC', 'Alternatives', portInput['BMA Asset Category'])
    portInput['BMA Asset Category'] = np. where(
            ((portInput['BMA Asset Category'] =='Bonds Cash and Govt') & (portInput['AIG Asset Class 3'] !="Cash") & (portInput['Mapped_BSCR_Rating']>2)), 
            "Bonds", portInput['BMA Asset Category'])
    portInput['BMA Asset Category'] = np. where(
            ((portInput['AIG Asset Class 3'] =="Derivative")), 
            "Bonds", portInput['BMA Asset Category'])
    
  # Combine ratings and asset class
    portInput['BMA_Category'] = portInput['BMA Asset Category']
    portInput['BMA_Category'] = np.where(
            portInput['Issuer Name'] =='LSTREET II, LLC', 'ML III', portInput['BMA_Category'])
    portInput['Mapped_BSCR_Rating'] = portInput['Mapped_BSCR_Rating'].astype(int)
    portInput['BMA_Category'] = np.where(
            ((portInput['BMA_Category'] == 'Bonds') | (portInput['BMA_Category'] == 'CMBS') | (portInput['BMA_Category'] == 'RMBS')),
            portInput['BMA_Category'] + "_" + portInput['Mapped_BSCR_Rating'].map(str), portInput['BMA_Category'])
 
    portInput = portInput.rename(columns = {'BMA Asset Category': 'BMA_Asset_Class'})

   # load ALBA duration and merge to portInput
    try:
        ALBA['dur'] = -ALBA['IR01']/ALBA['USD_MTM'] * 10000
        ALBA['POS_ID'] = ALBA['POS_ID'].astype(str)
        portInput['Lot Number DESC'] = portInput['Lot Number DESC'].astype(str)
        portInput = portInput.merge(ALBA[['POS_ID','dur']], how = 'left', left_on = 'Lot Number DESC', right_on = 'POS_ID').drop(columns = 'POS_ID')
        portInput['Effective Duration (WAMV)'] = portInput['Effective Duration (WAMV)'].fillna(portInput['dur'])
        portInput.drop(columns = 'dur')
        print('Check ALBA bucketed risk is loaded correctly. ' + str(ALBA['POS_ID'][0]) + '_duration is: ' + str(portInput[portInput['Lot Number DESC']==ALBA['POS_ID'][0]]['Effective Duration (WAMV)'].sum()) )
    except:
        print('ALBA bucketed risk is not loaded.')
    
    # Split out ML III Assets for concentration charge 
    portInput['Issuer Name'] = np.where(
            portInput['Issuer Name'] == 'LSTREET II, LLC', portInput['Issuer Name'] + '_' + portInput['Sec ID ID'].map(str), portInput['Issuer Name'])
       
    # Calculate cusip level charge
    portInput = portInput.merge(RcMap, how='left', left_on=['BMA_Category'],
                                right_on=['BMA_Category'])
    
    portInput['AssetCharge_Current'] = portInput['Market Value USD GAAP'] * portInput.Risk_Charge
    portInput['AssetCharge_Future'] = portInput['Market Value USD GAAP'] * portInput.Risk_Charge
    
    portInput['ConCharge_Current'] = portInput['Market Value USD GAAP'] * portInput.Concentration_Charge
    portInput['ConCharge_Future'] = portInput['Market Value USD GAAP'] * portInput.Concentration_Charge
  
    portInput['mv_dur']=portInput['Market Value USD GAAP'] * portInput['Effective Duration (WAMV)']
    
    # Existing Asset Charge    
    portInput['FIIndicator'] = portInput.BMA_Category.apply(
               lambda x: (0 if (x == 'Alternatives' or x == 'ML III' ) else 1))        
    portInput['EquityIndicator'] = portInput.BMA_Category.apply(
               lambda x: (1 if (x == 'Alternatives' or x == 'ML III') else 0))  
    
    portInput['FI Risk'] = 0
    portInput['FI Risk'] = np.where(
            portInput['FIIndicator'] == 1, portInput['AssetCharge_Current'], portInput['FI Risk'])
    portInput['Eq Risk_Current'] = 0
    portInput['Eq Risk_Current'] = np.where(
            portInput['EquityIndicator'] == 1, portInput['AssetCharge_Current'], portInput['Eq Risk_Current'])
    portInput['Eq Risk_Future'] = 0
    portInput['Eq Risk_Future'] = np.where(
            portInput['EquityIndicator'] == 1, portInput['AssetCharge_Future'], portInput['Eq Risk_Future'])
    
    portInput = portInput.drop_duplicates() 
    portInput.fillna(0, inplace=True)
    if output == 1:
        out_file = fileName + "_summary_test.xlsx" ### Vincent update 05/27/2019
        assetSummary = pd.ExcelWriter(out_file)
        portInput.to_excel(assetSummary, sheet_name='AssetSummaryFromPython', index=True, merge_cells=False)
        assetSummary.save()

    os.chdir(curr_dir)
    return portInput

#    ### === Prepare Asset Summary File === ###
#
#    ### This will distort BSCR calcualtion in 'Actual' ###
#    # Default setting NA 'AIG Derived Rating' to 'BBB'
#    portInput['AIG Derived Rating'].fillna('NA-BBB', inplace=True)
#    
#    portInput['AIG Derived Rating Update'] = portInput['AIG Derived Rating']
#
#    portInput['AIG Derived Rating Update'] = np.where(((portInput['AIG Asset Class 3'] == 'CMBS Agency') | (
#            portInput['AIG Asset Class 3'] == 'CMBS Non-Agency') | (portInput['AIG Asset Class 3'] == 'RMBS Agency') | (
#                                                               portInput['AIG Asset Class 3'] == 'RMBS Non-Agency')),
#                                                      portInput['Derived Rating Modified'], portInput['AIG Derived Rating'])
#    
#    # temporarily write to file for validation purpose
#    # portWriter = pd.ExcelWriter('./output/FRL_0228.xlsx')
#    # portInput.to_excel(portWriter, sheet_name='asset_port', index=False)
#    # portWriter.save()
#
#    groupByFields = ['Category', 'AIG Asset Class 3', 'AIG Derived Rating Update']
#
#    reportFields = {"YTW": "Market Value USD GAAP", \
#                    "OAS": "Market Value USD GAAP", \
#                    "Effective Duration (WAMV)": "Market Value USD GAAP", \
#                    "Effective Convexity": "Market Value USD GAAP", \
#                    "Spread Duration": "Market Value USD GAAP", \
#                    "Spread Convexity": "Market Value USD GAAP", \
#                    "WAL": "Quantity", \
#                    "KRD 1M": "Market Value USD GAAP", \
#                    "KRD 2M": "Market Value USD GAAP", \
#                    "KRD 3M": "Market Value USD GAAP", \
#                    "KRD 6M": "Market Value USD GAAP", \
#                    "KRD 9M": "Market Value USD GAAP", \
#                    "KRD 1Y": "Market Value USD GAAP", \
#                    "KRD 2Y": "Market Value USD GAAP", \
#                    "KRD 3Y": "Market Value USD GAAP", \
#                    "KRD 5Y": "Market Value USD GAAP", \
#                    "KRD 7Y": "Market Value USD GAAP", \
#                    "KRD 10Y": "Market Value USD GAAP", \
#                    "KRD 15Y": "Market Value USD GAAP", \
#                    "KRD 20Y": "Market Value USD GAAP", \
#                    "KRD 30Y": "Market Value USD GAAP"}
#
#    sumFields = groupByFields.copy()
#    sumFields.append('Market Value USD GAAP')
#    mv = portInput.loc[pd.notnull(portInput['Market Value USD GAAP']), sumFields].groupby(groupByFields).sum()
#
#    sumFields = groupByFields.copy()
#    sumFields.append('Market Value with Accrued Int USD GAAP')
#    mv_acc = portInput.loc[pd.notnull(portInput['Market Value with Accrued Int USD GAAP']), sumFields].groupby(
#        groupByFields).sum()
#
#    sumFields = groupByFields.copy()
#    sumFields.append('Book Value USD STAT')
#    bv = portInput.loc[pd.notnull(portInput['Book Value USD STAT']), sumFields].groupby(groupByFields).sum()
#
#    sumFields = groupByFields.copy()
#    sumFields.append('Book Value With Accrued Int USD STAT')
#    bv_acc = portInput.loc[pd.notnull(portInput['Book Value With Accrued Int USD STAT']), sumFields].groupby(
#        groupByFields).sum()
#
#    merge = pd.merge(left=mv, left_index=True, right=mv_acc, right_index=True, how='outer')
#    merge = pd.merge(left=merge, left_index=True, right=bv, right_index=True, how='outer')
#    merge = pd.merge(left=merge, left_index=True, right=bv_acc, right_index=True, how='outer')
#
#    for key, val in reportFields.items():
#        ds = portInput[pd.notnull(portInput[key])].groupby(groupByFields).apply(wavg(key, val))
#        df = pd.DataFrame(data=ds, columns=[key])
#        merge = pd.merge(left=merge, left_index=True, right=df, right_index=True, how='outer')
#
#    asset_count = merge.shape[0]
#    cal_category    = []
#    cal_asset_class = []
#    cal_rating      = []
#    
#    merge.fillna(0, inplace=True)
#    
#    for idx in range(0, asset_count, 1):
#        cal_category.append(merge.index[idx][0])
#        cal_asset_class.append(merge.index[idx][1])
#        cal_rating.append(merge.index[idx][2])
#    
#    merge['category_f']     = cal_category
#    merge['asset_class_f']  = cal_asset_class
#    merge['rating_f']       = cal_rating
#
#
##    Private Equity MV ajustment based on SPX
#    pe_return = IAL_App.eval_PE_return(eval_date, valDate_base)
#    mv_adj = []
#    mv_acc = []
#    for index, row in merge.iterrows():
#        if row['asset_class_f'] == "Private Equity Fund":
#            mv_adj.append(row['Market Value USD GAAP'] * (1 + pe_return) )
#            mv_acc
#        else:
#            mv_adj.append(row['Market Value USD GAAP'])
#
#    merge['mv_adj'] = mv_adj
#    
##   MV * Duration Calculations
#    merge['mv_dur']         = merge['mv_adj'] * merge['Effective Duration (WAMV)'] 
#    merge['acc_int']        = merge['Market Value with Accrued Int USD GAAP'] - merge['Market Value USD GAAP']
#    merge['mv_acc_int_adj'] = merge['mv_adj'] + merge['acc_int']
#    
#    if output == 1:
#        out_file = fileName[0:7] + "_summary.xlsx" ### fileName[0:25]
#        assetSummary = pd.ExcelWriter(out_file)
#        merge.to_excel(assetSummary, sheet_name='AssetSummaryFromPython', index=True, merge_cells=False)
#        assetSummary.save()
#        """
#        wb=xlw.Book.caller()
#        wb.sheets[0].range('A1').value = merge
#        wb.sheets[1].range('A1').value = portInput
#        """
##    return merge   
##       
##    if output == 1:
##        out_file = fileName[0:7] + "_summary_test.xlsx" ### Vincent update 05/27/2019
##        assetSummary = pd.ExcelWriter(out_file)
##        portInput.to_excel(assetSummary, sheet_name='AssetSummaryFromPython', index=True, merge_cells=False)
##        assetSummary.save()
#
#    return portInput, merge;
   
    
def stressed_actual_portfolio_feed(portInput, Scen, valDate, Asset_est):  
    ### @@@ to-do: mv_adj for dashboard for all shocks @@@ ###
    # MV = 'MV_USD_GAAP' for actual and mv_adj for estimate
    # calc_asset[MV] = .. .
    
    calc_asset = copy.deepcopy(portInput)
    calc_asset['Category'] = np.where((calc_asset['AIG Asset Class 3'] == "ML-III B-Notes"), "ML III", calc_asset['Category'])
    
    ### This process is NOT inclusive of shock on ALBA derivatives even under EBS reporting. Shock on ALBA and non-ALBA derivatives (duration & convexity are 0 in asset holdings file) will be read from Donovan's file in run_EBS.    
    ### @@@ to-do: check whether ALBA derivatives duration is 0 for dashboard asset holdings @@@ ### probably no as there is no ALBA bucketed risks file.. then where ALBA derivative dur comes from? ==> ALBA IR01
    
    # Asset_holding    Actual    Estimate
    # ALBA duration     Yes         No
    # ALBA mv           Yes         Yes?    
    # ALBA mv_dur       Yes         No
    
       
    # 1. Stressed Market Value - IR & Credit Spread Shocks
    ### Duration & convexity approach
    IR_shock = Scen['IR_Parallel_Shift_bps']/10000
    
    if Asset_est == 'Dur_Conv':
        print('Stress assets via [Dur_Conv]')        
        
        calc_asset['MV_USD_GAAP'] = np.where(  (calc_asset['FIIndicator'] == 1) & (calc_asset['Market Value with Accrued Int USD GAAP'] != 0) & (calc_asset['Category'] != 'ML III') & (calc_asset['AIG Asset Class 3'] != 'Derivative'),
                                                calc_asset['Market Value with Accrued Int USD GAAP'] * (1 - calc_asset['Spread Duration'] * calc_asset['Credit_Spread_Shock_bps']/10000 \
                                                                                                        + 1/2 * calc_asset['Spread Convexity'] * (IR_shock + calc_asset['Credit_Spread_Shock_bps']/10000) ** 2 * 100 \
                                                                                                        - calc_asset['Effective Duration (WAMV)'] * IR_shock) \
                                                - calc_asset['Accrued Int USD GAAP'],
                                                calc_asset['Market Value USD GAAP'] )
        
        calc_asset['Effective Duration (WAMV)'] = np.where( (calc_asset['FIIndicator'] == 1) & (calc_asset['Market Value with Accrued Int USD GAAP'] != 0) & (calc_asset['Category'] != 'ML III') & (calc_asset['AIG Asset Class 3'] != 'Derivative'),
                                                            calc_asset['Effective Duration (WAMV)'] - (100 * calc_asset['Effective Convexity'] - calc_asset['Effective Duration (WAMV)'] ** 2) * IR_shock \
                                                                                                    - (100 * calc_asset['Spread Convexity'] - calc_asset['Spread Duration'] ** 2) * calc_asset['Credit_Spread_Shock_bps']/10000,
                                                            calc_asset['Effective Duration (WAMV)'] )
        
    ### Bond Object Approach w/ US TSY curve
    elif Asset_est == 'Bond_Object':
        print('Stress assets via [Bond_Object]')
        base_curve  = IAL_App.createAkitZeroCurve(valDate, 'Treasury', 'USD')
        shock_curve = IAL_App.createAkitZeroCurve(valDate, 'Treasury', 'USD', IR_shift = Scen['IR_Parallel_Shift_bps'])
        
        calc_asset['Maturity_date']   = np.where( (calc_asset['Effective Duration (WAMV)'] != 0) & (calc_asset['YTW'] != 0),
                                                 calc_asset['WAL'].apply(lambda x: IAL.Util.addTerms(valDate, [str( math.ceil(max(1/12, x)*12) ) + "M"])[0] ),
                                                 datetime.datetime(2099, 12, 31, 0, 0))
                                               
        calc_asset['Bond_object']     = calc_asset.apply(lambda x: IAL.Bond.createBond(valDate, x['Maturity_date'], x['YTW']/100, 'S', '30/360', 0, 'U', 'F', 'None', 'None', x['Market Value USD GAAP']), axis = 1)
        
        calc_asset['OAS_bond_object'] = np.where( (calc_asset['Effective Duration (WAMV)'] != 0) & (calc_asset['YTW'] != 0),
                                                 calc_asset.apply(lambda x: 0 if x['Market Value USD GAAP'] < 0 else IAL.Bond.OAS(x['Bond_object'], valDate, x['Market Value USD GAAP'], base_curve), axis = 1),
                                                 0)
        
        calc_asset['MV_USD_GAAP']     = np.where( (calc_asset['Effective Duration (WAMV)'] != 0) & (calc_asset['YTW'] != 0) & (calc_asset['FIIndicator'] == 1) & (calc_asset['Market Value with Accrued Int USD GAAP'] != 0) & (calc_asset['Category'] != 'ML III') & (calc_asset['AIG Asset Class 3'] != 'Derivative'),
                                                 calc_asset.apply(lambda x: IAL.Bond.PVFromCurve(x['Bond_object'], valDate, shock_curve, (x['OAS_bond_object'] + x['Credit_Spread_Shock_bps'])/10000), axis = 1),
                                                 calc_asset['Market Value USD GAAP'])
        
        calc_asset['Effective Duration (WAMV)'] = np.where( (calc_asset['Category'] == 'ALBA') & (calc_asset['AIG Asset Class 3'] == 'Derivative'),
                                                            calc_asset['Effective Duration (WAMV)'],
                                                            calc_asset['Effective Duration (WAMV)'] - (100 * calc_asset['Effective Convexity'] - calc_asset['Effective Duration (WAMV)'] ** 2) * IR_shock \
                                                                                                    - (100 * calc_asset['Spread Convexity'] - calc_asset['Spread Duration'] ** 2) * calc_asset['Credit_Spread_Shock_bps']/10000
                                                            )
    
    # out_file = "Asset_bond_object_from_python.xlsx"
    # assetSummary = pd.ExcelWriter(out_file)
    # calc_asset.to_excel(assetSummary, sheet_name='AssetSummaryFromPython', index=True, merge_cells=False)
    # assetSummary.save()
            
    # ### === TEST === ###
    # calc_asset={'WAL': [6.0754669, 6.6724774, 0, 2.01014484767382],
    #             'YTW': [1.8718000/100, 4.0840000/100, 0.01, -1.5891/100],
    #             'MV' : [40865815.76,  27412719.21, 123456, 188155.64],
    #             'acc_int': [172849.47,  113097.51, 0, 0]
    #             }
    # calc_asset = pd.DataFrame(calc_asset)
        
    # calc_asset['maturity_Date'] = np.where( (calc_asset['WAL'] != 0),
    #                                         calc_asset['WAL'].apply(lambda x: IAL.Util.addTerms(valDate, [str( math.ceil(max(1/12, x)*12) ) + "M"])[0] ),
    #                                         datetime.datetime(2099, 12, 31, 0, 0))
    
    # calc_asset.dtypes
    
    
    # calc_asset['Bond_object']     = calc_asset.apply(lambda x: IAL.Bond.createBond(valDate, x['maturity_Date'], x['YTW'], 'S', '30/360', 0, 'U', 'F', 'None', 'None', x['MV']), axis = 1)
    # # calc_asset['OAS_bond_object'] = calc_asset.apply(lambda x: IAL.Bond.OAS(x['Bond_object'], valDate, x['MV'], base_curve), axis = 1)
    
    # # calc_asset['Bond_object'] = np.where( (calc_asset['WAL'] != 0),
    # #                                       calc_asset.apply(lambda x: IAL.Bond.createBond(valDate, x['maturity_Date'], x['YTW'], 'S', '30/360', 0, 'U', 'F', 'None', 'None', x['MV']), axis = 1),
    # #                                       'NA')
                                         
    # calc_asset['OAS_bond_object'] = np.where( (calc_asset['WAL'] != 0),
    #                                       calc_asset.apply(lambda x: 789 if x['MV'] < 0 else IAL.Bond.OAS(x['Bond_object'], valDate, x['MV'], base_curve), axis = 1),
    #                                       123122)
    
    # calc_asset['Price'] = np.where( (calc_asset['WAL'] != 0),
    #                                       calc_asset.apply(lambda x: IAL.Bond.PVFromCurve(x['Bond_object'], valDate, shock_curve, (x['OAS_bond_object']+166.25)/10000), axis = 1),
    #                                       0)
    
    # ### === TEST === ### 
    # WAL = 6.0754669
    # YTW = 1.8718000/100
    # MV = 40865815.76        
    # # acc_int = 172849.47
    
    # maturityDate = IAL.Util.addTerms(valDate, [str( math.ceil(max(1/12, WAL)*12) ) + "M"])[0]
    # base_curve  = IAL_App.createAkitZeroCurve(valDate, 'Treasury', 'USD')     
    # shock_curve = IAL_App.createAkitZeroCurve(valDate, 'Treasury', 'USD', IR_shift = -23.743476)   
                
    # Bond_object = IAL.Bond.createBond(valDate, maturityDate, YTW, 'S', '30/360', 0, 'U', 'F', 'None', 'None', MV) 
    
    # OAS = IAL.Bond.OAS(Bond_object, valDate, MV, base_curve)
    # print(OAS)

    # Price = IAL.Bond.PVFromCurve(Bond_object, valDate, shock_curve, (OAS+166.25)/10000) 
    # print(Price)
 
    
    # 2. Stressed MV * Dur        
    if Scen['IR_Parallel_Shift_bps'] != 0:
        calc_asset['mv_dur'] = np.where( (calc_asset['AIG Asset Class 3'] != 'Derivative'), # remove mv_dur for ALBA derivatives as its shock impact is quantified in [Load_stressed_derivatives_IR01] (it is 0 for non-ALBA derivatives anyway).
                                         calc_asset['MV_USD_GAAP'] * calc_asset['Effective Duration (WAMV)'], 0)
    else:
        calc_asset['mv_dur'] = calc_asset['MV_USD_GAAP'] * calc_asset['Effective Duration (WAMV)']
        
    # 3. Equity/Alts Shock
    Alt_shock = Scen['Alts_Retrun']
    alt_class = ['Common Equity', 'Private Equity Fund', 'Hedge Fund', 'Other Invested Assets']
    
    calc_asset['MV_USD_GAAP'] = np.where( ( np.isin(calc_asset['AIG Asset Class 3'], alt_class) ),
                                        calc_asset['Market Value with Accrued Int USD GAAP'] * (1 + Alt_shock) - calc_asset['Accrued Int USD GAAP'],
                                        calc_asset['MV_USD_GAAP'] )
          
    # 4. ML III shock
    MLIII_shock = Scen['MLIII_Return']
    
    calc_asset['MV_USD_GAAP'] = np.where( (calc_asset['Category'] == 'ML III'),
                                        calc_asset['Market Value with Accrued Int USD GAAP'] * (1 + MLIII_shock) - calc_asset['Accrued Int USD GAAP'],
                                        calc_asset['MV_USD_GAAP'] )
    
    # 5. Recalculate Asset Risk Charge (FI & Equity)
    calc_asset['AssetCharge_Current'] = calc_asset['MV_USD_GAAP'] * calc_asset.Risk_Charge
    calc_asset['AssetCharge_Future']  = calc_asset['MV_USD_GAAP'] * calc_asset.Risk_Charge
    
    calc_asset['FI Risk'] = np.where(calc_asset['FIIndicator'] == 1, calc_asset['AssetCharge_Current'], 0)
    
    calc_asset['Eq Risk_Current'] = np.where(calc_asset['EquityIndicator'] == 1, calc_asset['AssetCharge_Current'], 0)
    calc_asset['Eq Risk_Future']  = np.where(calc_asset['EquityIndicator'] == 1, calc_asset['AssetCharge_Future'],  0)
    
    # 6. Update stressed values in stressed asset holidng (so that BMA ALM BSCR shock can be applied on the stressed asset value when _stress_baseline == True)
    calc_asset['Market Value with Accrued Int USD GAAP'] = calc_asset['MV_USD_GAAP'] + calc_asset['Accrued Int USD GAAP']
    
    calc_asset['Market Value LCL GAAP'] = np.where((calc_asset['FX Rate LCL to USD STAT'] != 0) & (calc_asset['Market Value LCL GAAP'] != 0),
                                                   calc_asset['MV_USD_GAAP'] / calc_asset['FX Rate LCL to USD STAT'],
                                                   calc_asset['Market Value LCL GAAP'])
        
    # out_file = "Stressed_summary_" + Scen["Scen_Name"] + ".xlsx"
    # assetSummary = pd.ExcelWriter(out_file)
    # calc_asset.to_excel(assetSummary, sheet_name='AssetSummaryFromPython', index=True, merge_cells=False)
    # assetSummary.save()
    
    
    # a = EBS_Asset_Input_Base[EBS_Asset_Input_Base['FIIndicator'] == 1]['MV_USD_GAAP']-EBS_Asset_Input_Stressed[EBS_Asset_Input_Stressed['FIIndicator'] == 1]['MV_USD_GAAP']
    # a.sum()
   
    # b = []
    # b = EBS_Asset_Input_Stressed['Effective Duration (WAMV)']-EBS_Asset_Input_Base['Effective Duration (WAMV)']
    # b.sum()
    # EBS_Asset_Input_Base[(EBS_Asset_Input_Base['FIIndicator'] == 1)& (EBS_Asset_Input_Base['Market Value with Accrued Int USD GAAP'] != 0)&(EBS_Asset_Input_Base['Category'] != 'ML III')]['MV_USD_GAAP'].sum()

    # cusip_num = len(calc_asset)
    
    # for idx in range(0, cusip_num, 1):
    #     cals_cusip = calc_asset.iloc[idx]
        
    #     base_mv               = cals_cusip['Market Value with Accrued Int USD GAAP']
    #     base_dur              = cals_cusip['Effective Duration (WAMV)']
    #     base_convexity        = cals_cusip['Effective Convexity']
    #     base_spread_duration  = cals_cusip['Spread Duration']
    #     base_spread_convexity = cals_cusip['Spread Convexity']
            
    #     stressed_mv  = base_mv
    #     stressed_dur = base_dur
            
    #     if cals_cusip['FIIndicator'] == 1 and base_mv != 0 and cals_cusip['Category'] != 'ML III':   
               
    #         # 1. Stressed Market Value
    #         # 1.1 Credit spread shock                                                                 
    #         spread_shock = cals_cusip['Credit_Spread_Shock_bps'] / 10000
        
    #         each_change_in_asset = - base_mv * base_spread_duration * spread_shock \
    #                                 + base_mv * 1/2 * base_spread_convexity * spread_shock ** 2 * 100
            
    #         stressed_mv += each_change_in_asset ### spread impact
            
    #         # 1.2 IR shock                                                                      
    #         IR_shock = Scen['IR_Parallel_Shift_bps']/10000
                                   
    #         each_change_in_asset = - base_mv * base_dur * IR_shock \
    #                                 + base_mv * 1/2 * base_convexity * IR_shock ** 2 * 100 
                                      
    #         stressed_mv += each_change_in_asset ### IR impact
        
    #         stressed_mv -= cals_cusip['Accrued Int USD GAAP']
        
    #         # 2. Stressed Duration
    #         each_change_in_dur = - (100 * base_convexity - base_dur ** 2) * IR_shock \
    #                              - (100 * base_spread_convexity - base_spread_duration ** 2) * spread_shock
        
    #         stressed_dur += each_change_in_dur
        
    #     cals_cusip['Market Value USD GAAP']     = stressed_mv
    #     cals_cusip['Effective Duration (WAMV)'] = stressed_dur
        
    # 3. Stressed MV * Dur        
    # calc_asset['mv_dur'] = calc_asset['Market Value USD GAAP'] * calc_asset['Effective Duration (WAMV)']
    
    return calc_asset
